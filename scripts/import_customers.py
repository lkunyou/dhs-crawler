import pymysql
import openpyxl

DB_HOST = '8.163.58.109'
DB_PORT = 3306
DB_USER = 'thai_auto_parts_crm'
DB_PASSWORD = 'tDdY8NX2xJ6HpdHz'
DB_NAME = 'thai_auto_parts_crm'

def connect_db():
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset='utf8mb4'
    )

def import_customers(conn):
    cursor = conn.cursor()
    
    wb = openpyxl.load_workbook(r'E:\09.document\carparts\客户资料\泰国汽车配件客户开发清单_完整版.xlsx')
    ws = wb['客户清单总表']
    
    insert_count = 0
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=1).value
        if seq is None:
            continue
        
        company_name = ws.cell(row=row, column=2).value
        website = ws.cell(row=row, column=3).value
        address = ws.cell(row=row, column=4).value
        phone = ws.cell(row=row, column=5).value
        contact_email = ws.cell(row=row, column=6).value
        company_type = ws.cell(row=row, column=7).value
        main_products = ws.cell(row=row, column=8).value
        main_market = ws.cell(row=row, column=9).value
        company_size = ws.cell(row=row, column=10).value
        
        import_ability = ws.cell(row=row, column=11).value
        purchase_scale = ws.cell(row=row, column=12).value
        china_supplier_acceptance = ws.cell(row=row, column=13).value
        oem_aftermarket_match = ws.cell(row=row, column=14).value
        export_ability = ws.cell(row=row, column=15).value
        customization_match = ws.cell(row=row, column=16).value
        lead_score = ws.cell(row=row, column=17).value
        priority = ws.cell(row=row, column=18).value
        
        quality_requirement = ws.cell(row=row, column=19).value
        price_sensitivity = ws.cell(row=row, column=20).value
        delivery_requirement = ws.cell(row=row, column=21).value
        accept_china_factory = ws.cell(row=row, column=22).value
        customization_ability = ws.cell(row=row, column=23).value
        after_sales_requirement = ws.cell(row=row, column=24).value
        supply_chain_pain_points = ws.cell(row=row, column=25).value
        recommended_products = ws.cell(row=row, column=26).value
        recommended_channels = ws.cell(row=row, column=27).value
        first_email_strategy = ws.cell(row=row, column=28).value
        follow_up_immediately = ws.cell(row=row, column=29).value
        added_to_crm = ws.cell(row=row, column=30).value
        manual_follow_up = ws.cell(row=row, column=31).value
        remarketing = ws.cell(row=row, column=32).value
        next_follow_up_time = ws.cell(row=row, column=33).value
        development_email = ws.cell(row=row, column=34).value
        
        if company_name is None or company_name.strip() == '':
            continue
        
        company_type_enum = 'Importer'
        if company_type:
            if '批发' in company_type or '分销' in company_type:
                company_type_enum = 'Distributor'
            elif '进口' in company_type:
                company_type_enum = 'Importer'
            elif 'OEM' in company_type:
                company_type_enum = 'OEM'
            elif '零售' in company_type:
                company_type_enum = 'Retailer'
            elif '制造' in company_type:
                company_type_enum = 'Manufacturer'
            else:
                company_type_enum = 'Other'
        
        lead_grade = priority if priority in ['S', 'A', 'B', 'C'] else 'B'
        
        employee_count = ''
        if company_size:
            if '大' in company_size:
                employee_count = '500+'
            elif '中' in company_size:
                employee_count = '100-500'
            elif '小' in company_size:
                employee_count = '1-100'
        
        accept_china = '是' if accept_china_factory and '是' in str(accept_china_factory) else '否'
        
        try:
            sql = """
INSERT INTO p_company (company_name, website, address, phone, email, company_type, 
                       main_products, main_market, employee_count,
                       import_ability, purchase_scale, china_supplier_acceptance, 
                       oem_aftermarket_match, export_ability, customization_match,
                       lead_score, lead_grade,
                       quality_requirement, price_sensitivity, delivery_requirement,
                       accept_china_factory, customization_ability, after_sales_requirement,
                       supply_chain_pain_points, recommended_products, recommended_channels,
                       first_email_strategy, follow_up_immediately, added_to_crm,
                       manual_follow_up, remarketing, development_email,
                       country, status, source, is_auto_parts_core, is_importer_distributor)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            values = (
                company_name, website, address, phone, contact_email, company_type_enum,
                main_products, main_market, employee_count,
                import_ability, purchase_scale, china_supplier_acceptance,
                oem_aftermarket_match, export_ability, customization_match,
                lead_score, lead_grade,
                quality_requirement, price_sensitivity, delivery_requirement,
                accept_china, customization_ability, after_sales_requirement,
                supply_chain_pain_points, recommended_products, recommended_channels,
                first_email_strategy, follow_up_immediately, added_to_crm,
                manual_follow_up, remarketing, development_email,
                'Thailand', 'New', 'Manual', 1, 1
            )
            cursor.execute(sql, values)
            insert_count += 1
            if insert_count % 5 == 0:
                conn.commit()
                print(f"已导入 {insert_count} 条数据...")
        except Exception as e:
            print(f"❌ 导入失败 (行{row}): {company_name} - {e}")
    
    conn.commit()
    print(f"\n✅ 共导入 {insert_count} 条客户数据")
    cursor.close()

def verify_data(conn):
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) as total FROM p_company")
    total = cursor.fetchone()[0]
    print(f"\n📊 数据库中共有 {total} 条客户记录")
    
    cursor.execute("SELECT lead_grade, COUNT(*) as count FROM p_company GROUP BY lead_grade ORDER BY lead_grade")
    grades = cursor.fetchall()
    print("\n📈 客户等级分布:")
    for grade, count in grades:
        print(f"  {grade}: {count} 条")
    
    cursor.execute("SELECT company_type, COUNT(*) as count FROM p_company GROUP BY company_type")
    types = cursor.fetchall()
    print("\n🏢 公司类型分布:")
    for ct, count in types:
        print(f"  {ct}: {count} 条")
    
    cursor.execute("SELECT MIN(lead_score), MAX(lead_score), AVG(lead_score) FROM p_company")
    score_stats = cursor.fetchone()
    print(f"\n📊 评分统计:")
    print(f"  最低评分: {score_stats[0]}")
    print(f"  最高评分: {score_stats[1]}")
    print(f"  平均评分: {score_stats[2]:.1f}")
    
    cursor.close()

if __name__ == '__main__':
    print("=" * 60)
    print("正在连接数据库并导入客户数据...")
    print("=" * 60)
    
    try:
        conn = connect_db()
        print("✅ 数据库连接成功")
        
        import_customers(conn)
        verify_data(conn)
        
        conn.close()
        print("\n🎉 客户数据导入完成！")
        
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()