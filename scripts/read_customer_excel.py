import openpyxl

wb = openpyxl.load_workbook(r'E:\09.document\carparts\客户资料\泰国汽车配件客户开发清单_完整版.xlsx')

print("工作表列表:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

print("\n" + "=" * 60)
print("读取第一个工作表的结构...")
print("=" * 60)

ws = wb[wb.sheetnames[0]]

print(f"\n总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")

print("\n表头行(第1行):")
for col in range(1, ws.max_column + 1):
    value = ws.cell(row=1, column=col).value
    print(f"  列{col}: {value}")

print("\n前3行数据示例:")
for row in range(2, min(5, ws.max_row + 1)):
    print(f"\n--- 第{row}行 ---")
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        value = ws.cell(row=row, column=col).value
        print(f"  {header}: {value}")