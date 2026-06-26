import openpyxl

wb = openpyxl.load_workbook(r'E:\09.document\carparts\产品\产品价格2026-6-25(1).xlsx')
ws = wb['Sheet1']

rows = []
for row in range(4, ws.max_row + 1):
    seq = ws.cell(row=row, column=1).value
    if seq is None:
        continue
    
    product_name = ws.cell(row=row, column=2).value
    oe_no = ws.cell(row=row, column=3).value
    weight = ws.cell(row=row, column=4).value
    model = ws.cell(row=row, column=5).value
    dimensions = ws.cell(row=row, column=6).value
    pkg_length = ws.cell(row=row, column=7).value
    pkg_width = ws.cell(row=row, column=8).value
    pkg_height = ws.cell(row=row, column=9).value
    qty_per_pkg = ws.cell(row=row, column=10).value
    car_model = ws.cell(row=row, column=11).value
    production_date = ws.cell(row=row, column=12).value
    price_cny = ws.cell(row=row, column=13).value
    quote_price_cny = ws.cell(row=row, column=14).value
    
    if product_name is None or product_name.strip() == '':
        continue
    
    if isinstance(oe_no, str) and 'VLOOKUP' in oe_no:
        oe_no = None
    
    if isinstance(car_model, str) and 'VLOOKUP' in car_model:
        car_model = None
    
    if isinstance(production_date, str) and 'VLOOKUP' in production_date:
        production_date = None
    
    if price_cny and isinstance(price_cny, (int, float)):
        unit_price_usd = price_cny / 7.2
    else:
        unit_price_usd = 0
    
    category = 'exterior'
    if '扰流板' in product_name:
        category = 'exterior'
    elif '尾管' in product_name:
        category = 'exterior'
    elif '中网' in product_name:
        category = 'exterior'
    elif '前唇' in product_name:
        category = 'exterior'
    elif '侧裙' in product_name:
        category = 'exterior'
    elif '后唇' in product_name:
        category = 'exterior'
    elif '轮眉' in product_name:
        category = 'exterior'
    elif '行李架' in product_name:
        category = 'exterior'
    elif '踏板' in product_name:
        category = 'exterior'
    elif '护板' in product_name:
        category = 'exterior'
    elif '发动机' in product_name:
        category = 'engine'
    elif '刹车' in product_name:
        category = 'brake'
    elif '悬挂' in product_name:
        category = 'suspension'
    elif '电气' in product_name or '灯' in product_name:
        category = 'electrical'
    else:
        category = 'exterior'
    
    spec_parts = []
    if model:
        spec_parts.append(f"型号: {model}")
    if dimensions:
        spec_parts.append(f"尺寸: {dimensions}")
    if weight and isinstance(weight, (int, float)):
        spec_parts.append(f"重量: {weight}g")
    if car_model:
        spec_parts.append(f"适用车型: {car_model}")
    if pkg_length and pkg_width and pkg_height:
        spec_parts.append(f"包装: {pkg_length}x{pkg_width}x{pkg_height}mm")
    if qty_per_pkg:
        spec_parts.append(f"每箱: {qty_per_pkg}件")
    
    specification = '; '.join(spec_parts)
    
    rows.append({
        'product_name': product_name,
        'product_code': oe_no,
        'category': category,
        'brand': 'OEM',
        'specification': specification,
        'unit_price': round(unit_price_usd, 2),
        'stock': 100,
        'description': f"{product_name} - 适用车型: {car_model or '通用'}",
        'status': 'active'
    })

print(f"共读取 {len(rows)} 条产品数据")
print()

sql_lines = []
for i, row in enumerate(rows, 1):
    sql_lines.append(f"""INSERT INTO p_product (product_name, product_code, category, brand, specification, unit_price, stock, description, status) VALUES (
    '{row['product_name'].replace("'", "''")}',
    {'NULL' if not row['product_code'] else f"'{row['product_code']}'"},
    '{row['category']}',
    '{row['brand']}',
    '{row['specification'].replace("'", "''")}',
    {row['unit_price']},
    {row['stock']},
    '{row['description'].replace("'", "''")}',
    'active'
);""")

sql_content = '\n\n'.join(sql_lines)
sql_file = r'/database/import_products.sql'
with open(sql_file, 'w', encoding='utf-8') as f:
    f.write(sql_content)

print(f"SQL文件已生成: {sql_file}")
print(f"共 {len(rows)} 条插入语句")