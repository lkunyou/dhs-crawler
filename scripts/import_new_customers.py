import pymysql
import openpyxl
import re

DB_HOST = '8.163.58.109'
DB_PORT = 3306
DB_USER = 'thai_auto_parts_crm'
DB_PASSWORD = 'tDdY8NX2xJ6HpdHz'
DB_NAME = 'thai_auto_parts_crm'

def connect_db():
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset='utf8mb4'
    )

def parse_revenue(revenue_str):
    if not revenue_str:
        return 0
    
    revenue_str = str(revenue_str).strip()
    
    match = re.search(r'([\d.,]+)', revenue_str)
    if not match:
        return 0
    
    num_str = match.group(1).replace(',', '')
    
    try:
        revenue = float(num_str)
        
        if '万' in revenue_str:
            revenue *= 10000
        
        return revenue
    except:
        return 0

def parse_priority(priority_str):
    if not priority_str:
        return 'B', 60
    
    stars = str(priority_str).count('★')
    
    if stars >= 5:
        return 'S', 90
    elif stars == 4:
        return 'A', 75
    elif stars == 3:
        return 'B', 60
    elif stars == 2:
        return 'C', 40
    else:
        return 'C', 30

def parse_potential(potential_str):
    if not potential_str:
        return 15
    
    potential_str = str(potential_str)
    
    if '极高' in potential_str:
        return 25
    elif '高' in potential_str:
        return 20
    elif '中' in potential_str:
        return 12
    elif '低' in potential_str:
        return 5
    else:
        return 10

def extract_email_from_contact(contact_str):
    if not contact_str:
        return None
    
    match = re.search(r'[\w.-]+@[\w.-]+\.\w+', str(contact_str))
    return match.group(0) if match else None

def extract_contact_name(contact_str):
    if not contact_str:
        return None
    
    contact_str = str(contact_str)
    
    name_parts = []
    parts = contact_str.split()
    
    for part in parts:
        if part.startswith('Mr.') or part.startswith('Ms.') or part.startswith('Mrs.'):
            if len(parts) > parts.index(part) + 1:
                name_parts.append(part + ' ' + parts[parts.index(part) + 1])
        elif '@' not in part and len(part) > 2:
            name_parts.append(part)
    
    if name_parts:
        return ' '.join(name_parts)[:100]
    
    return contact_str[:100]

def import_customers(conn):
    cursor = conn.cursor()
    
    wb = openpyxl.load_workbook(r'E:\09.document\carparts\客户资料\客户资料.xlsx')
    ws = wb['Sheet1']
    
    inserted_count = 0
    updated_count = 0
    
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=1).value
        if seq is None:
            continue
        
        company_name = ws.cell(row=row, column=2).value
        website = ws.cell(row=row, column=3).value
        address = ws.cell(row=row, column=4).value
        phone = ws.cell(row=row, column=5).value
        contact_info = ws.cell(row=row, column=6).value
        main_products = ws.cell(row=row, column=7).value
        main_market = ws.cell(row=row, column=8).value
        revenue = ws.cell(row=row, column=9).value
        purchase_potential = ws.cell(row=row, column=10).value
        priority = ws.cell(row=row, column=11).value
        
        if company_name is None or company_name.strip() == '':
            continue
        
        cursor.execute("SELECT id FROM p_company WHERE company_name = %s", (company_name,))
        existing = cursor.fetchone()
        
        email = extract_email_from_contact(contact_info)
        contact_name = extract_contact_name(contact_info)
        
        revenue_value = parse_revenue(revenue)
        lead_grade, lead_score = parse_priority(priority)
        purchase_scale = parse_potential(purchase_potential)
        
        import_ability = 25 if revenue_value >= 5000000 else 20 if revenue_value >= 1000000 else 15
        
        if '出口' in str(main_market) or 'export' in str(main_market).lower():
            export_ability = 9
        else:
            export_ability = 5
        
        china_supplier_acceptance = 18
        oem_aftermarket_match = 14
        customization_match = 4
        
        if existing:
            company_id = existing[0]
            
            update_sql = """
UPDATE p_company SET 
    website = %s, address = %s, phone = %s, email = %s,
    main_products = %s, main_market = %s,
    lead_score = %s, lead_grade = %s,
    import_ability = %s, purchase_scale = %s,
    china_supplier_acceptance = %s, oem_aftermarket_match = %s,
    export_ability = %s, customization_match = %s,
    status = %s, source = %s,
    is_auto_parts_core = %s, is_importer_distributor = %s
WHERE id = %s
            """
            cursor.execute(update_sql, (
                website, address, phone, email,
                main_products, main_market,
                lead_score, lead_grade,
                import_ability, purchase_scale,
                china_supplier_acceptance, oem_aftermarket_match,
                export_ability, customization_match,
                'New', 'Manual',
                1, 1,
                company_id
            ))
            updated_count += 1
        
        else:
            insert_sql = """
INSERT INTO p_company (
    company_name, website, address, phone, email,
    main_products, main_market,
    lead_score, lead_grade,
    import_ability, purchase_scale,
    china_supplier_acceptance, oem_aftermarket_match,
    export_ability, customization_match,
    country, status, source,
    is_auto_parts_core, is_importer_distributor
) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(insert_sql, (
                company_name, website, address, phone, email,
                main_products, main_market,
                lead_score, lead_grade,
                import_ability, purchase_scale,
                china_supplier_acceptance, oem_aftermarket_match,
                export_ability, customization_match,
                'Thailand', 'New', 'Manual',
                1, 1
            ))
            inserted_count += 1
        
        if (inserted_count + updated_count) % 10 == 0:
            conn.commit()
            print(f"已处理 {inserted_count + updated_count} 条数据...")
    
    conn.commit()
    print(f"\n✅ 共处理 {inserted_count + updated_count} 条客户数据")
    print(f"   新增: {inserted_count} 条")
    print(f"   更新: {updated_count} 条")
    cursor.close()

def verify_data(conn):
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) as total FROM p_company")
    total = cursor.fetchone()[0]
    print(f"\n📊 数据库中共有 {total} 条客户记录")
    
    cursor.execute("SELECT lead_grade, COUNT(*) as count FROM p_company GROUP BY lead_grade ORDER BY lead_grade")
    grades = cursor.fetchall()
    print("\n📈 客户等级分布:")
    for grade, count in grades:
        print(f"  {grade}: {count} 条")
    
    cursor.execute("SELECT COUNT(*) FROM p_company WHERE purchase_scale >= 20")
    high_potential = cursor.fetchone()[0]
    print(f"\n💰 高采购潜力客户: {high_potential} 条")
    
    cursor.execute("SELECT COUNT(*) FROM p_company WHERE export_ability >= 8")
    export_companies = cursor.fetchone()[0]
    print(f"🌍 有出口能力的客户: {export_companies} 条")
    
    cursor.close()

if __name__ == '__main__':
    print("=" * 60)
    print("正在连接数据库并导入客户数据...")
    print("=" * 60)
    
    try:
        conn = connect_db()
        print("✅ 数据库连接成功")
        
        import_customers(conn)
        verify_data(conn)
        
        conn.close()
        print("\n🎉 客户数据导入完成！")
        
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()