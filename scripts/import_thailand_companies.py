import pymysql
import openpyxl

DB_HOST = '8.163.58.109'
DB_PORT = 3306
DB_USER = 'thai_auto_parts_crm'
DB_PASSWORD = 'tDdY8NX2xJ6HpdHz'
DB_NAME = 'thai_auto_parts_crm'

def connect_db():
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset='utf8mb4'
    )

def parse_category(category):
    if not category:
        return 'Other'
    category = str(category).strip()
    if 'Manufacturer' in category or 'manufacturer' in category:
        return 'Manufacturer'
    elif 'Wholesale' in category or 'wholesale' in category:
        if 'Retail' in category or 'retail' in category:
            return 'Distributor'
        return 'Distributor'
    elif 'Retail' in category or 'retail' in category:
        return 'Retailer'
    elif 'OEM' in category or 'oem' in category:
        return 'OEM'
    elif 'Import' in category or 'import' in category:
        return 'Importer'
    return 'Distributor'

def extract_company_name(name):
    if not name:
        return None
    name = str(name).strip()
    if '(' in name and ')' in name:
        start = name.find('(')
        end = name.rfind(')')
        if end > start:
            eng_name = name[start+1:end].strip()
            th_name = name[:start].strip()
            return eng_name, th_name
    return name, None

def import_customers(conn):
    cursor = conn.cursor()
    
    wb = openpyxl.load_workbook('thailand_auto_parts_companies.xlsx')
    ws = wb['Sheet1']
    
    inserted_count = 0
    skipped_count = 0
    
    for row in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=1).value
        if raw_name is None or str(raw_name).strip() == '':
            skipped_count += 1
            continue
        
        category = ws.cell(row=row, column=2).value
        address = ws.cell(row=row, column=3).value
        phone = ws.cell(row=row, column=4).value
        email = ws.cell(row=row, column=5).value
        website = ws.cell(row=row, column=6).value
        description = ws.cell(row=row, column=7).value
        
        company_name_en, company_name_th = extract_company_name(raw_name)
        company_type = parse_category(category)
        
        if company_name_en == 'Thailand':
            skipped_count += 1
            continue
        
        main_products = description if description else ''
        
        cursor.execute("SELECT id FROM p_company WHERE company_name = %s", (company_name_en,))
        existing = cursor.fetchone()
        
        if existing:
            print(f"⚠️  跳过重复: {company_name_en}")
            skipped_count += 1
            continue
        
        try:
            sql = """
INSERT INTO p_company (
    company_name, company_name_th, company_name_en,
    company_type, website, address, phone, email,
    main_products, lead_score, lead_grade,
    country, status, source,
    is_auto_parts_core, is_importer_distributor,
    employee_count, import_ability, purchase_scale,
    china_supplier_acceptance, oem_aftermarket_match,
    export_ability, customization_match
) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            values = (
                company_name_en, company_name_th, company_name_en,
                company_type, website, address, phone, email,
                main_products, 50, 'B',
                'Thailand', 'New', 'Manual',
                1, 1 if company_type in ['Importer', 'Distributor'] else 0,
                '', 15, 15,
                15, 10, 5, 5
            )
            cursor.execute(sql, values)
            inserted_count += 1
            
            if inserted_count % 10 == 0:
                conn.commit()
                print(f"已导入 {inserted_count} 条数据...")
                
        except Exception as e:
            print(f"❌ 导入失败 (行{row}): {company_name_en} - {e}")
            skipped_count += 1
    
    conn.commit()
    print(f"\n✅ 共导入 {inserted_count} 条客户数据")
    print(f"⚠️  跳过 {skipped_count} 条数据")
    cursor.close()

def verify_data(conn):
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) as total FROM p_company")
    total = cursor.fetchone()[0]
    print(f"\n📊 数据库中共有 {total} 条客户记录")
    
    cursor.execute("SELECT lead_grade, COUNT(*) as count FROM p_company GROUP BY lead_grade ORDER BY lead_grade")
    grades = cursor.fetchall()
    print("\n📈 客户等级分布:")
    for grade, count in grades:
        print(f"  {grade}: {count} 条")
    
    cursor.execute("SELECT company_type, COUNT(*) as count FROM p_company GROUP BY company_type")
    types = cursor.fetchall()
    print("\n🏢 公司类型分布:")
    for ct, count in types:
        print(f"  {ct}: {count} 条")
    
    cursor.close()

if __name__ == '__main__':
    print("=" * 60)
    print("正在导入 thailand_auto_parts_companies.xlsx 到客户表...")
    print("=" * 60)
    
    try:
        conn = connect_db()
        print("✅ 数据库连接成功")
        
        import_customers(conn)
        verify_data(conn)
        
        conn.close()
        print("\n🎉 客户数据导入完成！")
        
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()