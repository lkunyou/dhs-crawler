import pymysql
import openpyxl
import re

DB_HOST = '8.163.58.109'
DB_PORT = 3306
DB_USER = 'thai_auto_parts_crm'
DB_PASSWORD = 'tDdY8NX2xJ6HpdHz'
DB_NAME = 'thai_auto_parts_crm'

def connect_db():
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset='utf8mb4'
    )

def extract_email(contact_str):
    if not contact_str:
        return None
    match = re.search(r'[\w.-]+@[\w.-]+\.\w+', str(contact_str))
    return match.group(0) if match else None

def extract_contact_name(contact_str):
    if not contact_str:
        return None
    contact_str = str(contact_str)
    if '@' in contact_str:
        contact_str = contact_str.split('@')[0]
    return contact_str.strip()[:100]

def parse_company_type(type_str):
    if not type_str:
        return 'Distributor'
    type_str = str(type_str)
    if '进口' in type_str:
        return 'Importer'
    if '分销' in type_str or '批发' in type_str:
        return 'Distributor'
    if '制造' in type_str:
        return 'Manufacturer'
    if '零售' in type_str:
        return 'Retailer'
    return 'Distributor'

def parse_priority(priority_str):
    if not priority_str:
        return 'B', 60
    priority_str = str(priority_str).strip().upper()
    if priority_str == 'S' or priority_str == '★★★★★':
        return 'S', 90
    elif priority_str == 'A':
        return 'A', 75
    elif priority_str == 'B':
        return 'B', 60
    elif priority_str == 'C':
        return 'C', 40
    else:
        return 'B', 60

def parse_next_follow_up(days_str):
    if not days_str:
        return None
    days_str = str(days_str)
    match = re.search(r'(\d+)', days_str)
    if match:
        return int(match.group(1))
    return None

def import_customers(conn):
    cursor = conn.cursor()
    
    wb = openpyxl.load_workbook(r'E:\09.document\carparts\客户资料\泰国汽车配件客户开发清单_完整版.xlsx')
    ws = wb[wb.sheetnames[0]]
    
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers[header.strip()] = col
    
    print(f"解析到 {len(headers)} 个字段:")
    for h in headers:
        print(f"  - {h}")
    
    inserted_count = 0
    updated_count = 0
    
    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=headers.get('序号', 1)).value
        if seq is None:
            continue
        
        company_name = ws.cell(row=row, column=headers.get('公司名称', 2)).value
        if not company_name:
            continue
        
        website = ws.cell(row=row, column=headers.get('官网', 3)).value
        address = ws.cell(row=row, column=headers.get('地址', 4)).value
        phone = ws.cell(row=row, column=headers.get('联系电话', 5)).value
        contact_info = ws.cell(row=row, column=headers.get('核心联系人/邮箱', 6)).value
        company_type = ws.cell(row=row, column=headers.get('公司类型', 7)).value
        main_products = ws.cell(row=row, column=headers.get('主营产品', 8)).value
        main_market = ws.cell(row=row, column=headers.get('主要销售市场', 9)).value
        
        import_ability = ws.cell(row=row, column=headers.get('进口能力(0-25)', 11)).value
        purchase_scale = ws.cell(row=row, column=headers.get('采购规模(0-25)', 12)).value
        china_supplier_acceptance = ws.cell(row=row, column=headers.get('中国供应商接受度(0-20)', 13)).value
        oem_aftermarket_match = ws.cell(row=row, column=headers.get('OEM/Aftermarket匹配度(0-15)', 14)).value
        export_ability = ws.cell(row=row, column=headers.get('出口能力(0-10)', 15)).value
        customization_match = ws.cell(row=row, column=headers.get('定制需求匹配度(0-5)', 16)).value
        lead_score = ws.cell(row=row, column=headers.get('总评分', 17)).value
        priority = ws.cell(row=row, column=headers.get('开发优先级', 18)).value
        
        quality_requirement = ws.cell(row=row, column=headers.get('质量要求', 19)).value
        price_sensitivity = ws.cell(row=row, column=headers.get('价格敏感度', 20)).value
        delivery_requirement = ws.cell(row=row, column=headers.get('交期要求', 21)).value
        accept_china_factory = ws.cell(row=row, column=headers.get('接受中国工厂', 22)).value
        customization_ability = ws.cell(row=row, column=headers.get('定制能力需求', 23)).value
        after_sales_requirement = ws.cell(row=row, column=headers.get('售后要求', 24)).value
        supply_chain_pain_points = ws.cell(row=row, column=headers.get('当前供应链痛点', 25)).value
        recommended_products = ws.cell(row=row, column=headers.get('推荐切入产品', 26)).value
        recommended_channels = ws.cell(row=row, column=headers.get('推荐渠道', 27)).value
        first_email_strategy = ws.cell(row=row, column=headers.get('首封开发信策略', 28)).value
        
        follow_up_immediately = ws.cell(row=row, column=headers.get('是否立即跟进', 29)).value
        added_to_crm = ws.cell(row=row, column=headers.get('加入CRM', 30)).value
        manual_follow_up = ws.cell(row=row, column=headers.get('人工跟进', 31)).value
        remarketing = ws.cell(row=row, column=headers.get('再营销', 32)).value
        development_email = ws.cell(row=row, column=headers.get('开发信(英文)', 34)).value
        
        email = extract_email(contact_info)
        core_contact = extract_contact_name(contact_info)
        company_type_en = parse_company_type(company_type)
        lead_grade, default_score = parse_priority(priority)
        
        if lead_score is None:
            lead_score = default_score
        
        cursor.execute("SELECT id FROM p_company WHERE company_name = %s", (company_name,))
        existing = cursor.fetchone()
        
        if existing:
            company_id = existing[0]
            
            update_sql = """
UPDATE p_company SET 
    website = IFNULL(%s, website),
    address = IFNULL(%s, address),
    phone = IFNULL(%s, phone),
    email = IFNULL(%s, email),
    company_type = IFNULL(%s, company_type),
    main_products = IFNULL(%s, main_products),
    main_market = IFNULL(%s, main_market),
    import_ability = IFNULL(%s, import_ability),
    purchase_scale = IFNULL(%s, purchase_scale),
    china_supplier_acceptance = IFNULL(%s, china_supplier_acceptance),
    oem_aftermarket_match = IFNULL(%s, oem_aftermarket_match),
    export_ability = IFNULL(%s, export_ability),
    customization_match = IFNULL(%s, customization_match),
    lead_score = %s,
    lead_grade = %s,
    quality_requirement = IFNULL(%s, quality_requirement),
    price_sensitivity = IFNULL(%s, price_sensitivity),
    delivery_requirement = IFNULL(%s, delivery_requirement),
    accept_china_factory = IFNULL(%s, accept_china_factory),
    customization_ability = IFNULL(%s, customization_ability),
    after_sales_requirement = IFNULL(%s, after_sales_requirement),
    supply_chain_pain_points = IFNULL(%s, supply_chain_pain_points),
    recommended_products = IFNULL(%s, recommended_products),
    recommended_channels = IFNULL(%s, recommended_channels),
    first_email_strategy = IFNULL(%s, first_email_strategy),
    follow_up_immediately = IFNULL(%s, follow_up_immediately),
    added_to_crm = IFNULL(%s, added_to_crm),
    manual_follow_up = IFNULL(%s, manual_follow_up),
    remarketing = IFNULL(%s, remarketing),
    development_email = IFNULL(%s, development_email),
    core_contact = IFNULL(%s, core_contact),
    status = 'New',
    source = 'Manual',
    is_auto_parts_core = 1,
    is_importer_distributor = %s
WHERE id = %s
            """
            
            is_importer = 1 if '进口' in str(company_type) or '分销' in str(company_type) else 0
            
            cursor.execute(update_sql, (
                website, address, phone, email,
                company_type_en, main_products, main_market,
                import_ability, purchase_scale, china_supplier_acceptance,
                oem_aftermarket_match, export_ability, customization_match,
                lead_score, lead_grade,
                quality_requirement, price_sensitivity, delivery_requirement,
                accept_china_factory, customization_ability, after_sales_requirement,
                supply_chain_pain_points, recommended_products, recommended_channels,
                first_email_strategy, follow_up_immediately, added_to_crm,
                manual_follow_up, remarketing, development_email,
                core_contact,
                is_importer,
                company_id
            ))
            updated_count += 1
        
        else:
            is_importer = 1 if '进口' in str(company_type) or '分销' in str(company_type) else 0
            
            insert_sql = """
INSERT INTO p_company (
    company_name, website, address, phone, email,
    company_type, main_products, main_market,
    import_ability, purchase_scale, china_supplier_acceptance,
    oem_aftermarket_match, export_ability, customization_match,
    lead_score, lead_grade,
    quality_requirement, price_sensitivity, delivery_requirement,
    accept_china_factory, customization_ability, after_sales_requirement,
    supply_chain_pain_points, recommended_products, recommended_channels,
    first_email_strategy, follow_up_immediately, added_to_crm,
    manual_follow_up, remarketing, development_email,
    core_contact,
    country, status, source,
    is_auto_parts_core, is_importer_distributor
) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            cursor.execute(insert_sql, (
                company_name, website, address, phone, email,
                company_type_en, main_products, main_market,
                import_ability, purchase_scale, china_supplier_acceptance,
                oem_aftermarket_match, export_ability, customization_match,
                lead_score, lead_grade,
                quality_requirement, price_sensitivity, delivery_requirement,
                accept_china_factory, customization_ability, after_sales_requirement,
                supply_chain_pain_points, recommended_products, recommended_channels,
                first_email_strategy, follow_up_immediately, added_to_crm,
                manual_follow_up, remarketing, development_email,
                core_contact,
                'Thailand', 'New', 'Manual',
                1, is_importer
            ))
            inserted_count += 1
        
        if (inserted_count + updated_count) % 5 == 0:
            conn.commit()
            print(f"已处理 {inserted_count + updated_count} 条数据...")
    
    conn.commit()
    print(f"\n✅ 共处理 {inserted_count + updated_count} 条客户数据")
    print(f"   新增: {inserted_count} 条")
    print(f"   更新: {updated_count} 条")
    cursor.close()

def verify_data(conn):
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) as total FROM p_company")
    total = cursor.fetchone()[0]
    print(f"\n📊 数据库中共有 {total} 条客户记录")
    
    cursor.execute("SELECT lead_grade, COUNT(*) as count FROM p_company GROUP BY lead_grade ORDER BY lead_grade")
    grades = cursor.fetchall()
    print("\n📈 客户等级分布:")
    for grade, count in grades:
        print(f"  {grade}: {count} 条")
    
    cursor.execute("SELECT COUNT(*) FROM p_company WHERE first_email_strategy IS NOT NULL AND first_email_strategy != ''")
    strategy_count = cursor.fetchone()[0]
    print(f"\n📝 有首封开发信策略的客户: {strategy_count} 条")
    
    cursor.execute("SELECT COUNT(*) FROM p_company WHERE development_email IS NOT NULL AND development_email != ''")
    email_count = cursor.fetchone()[0]
    print(f"📧 有开发信内容的客户: {email_count} 条")
    
    cursor.execute("SELECT COUNT(*) FROM p_company WHERE company_type IS NOT NULL AND company_type != ''")
    type_count = cursor.fetchone()[0]
    print(f"🏢 有公司类型的客户: {type_count} 条")
    
    cursor.execute("""
SELECT company_name, lead_score, lead_grade, company_type
FROM p_company 
WHERE lead_score >= 85
ORDER BY lead_score DESC
LIMIT 5
""")
    results = cursor.fetchall()
    print(f"\n🏆 高评分客户示例:")
    for row in results:
        print(f"  {row[0][:40]:<40} | 评分: {row[1]:<3} | 等级: {row[2]} | 类型: {row[3]}")
    
    cursor.close()

if __name__ == '__main__':
    print("=" * 60)
    print("导入完整版客户开发清单到数据库")
    print("=" * 60)
    
    try:
        conn = connect_db()
        print("✅ 数据库连接成功")
        
        import_customers(conn)
        verify_data(conn)
        
        conn.close()
        print("\n🎉 客户数据导入完成！")
        
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()