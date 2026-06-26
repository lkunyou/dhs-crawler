import pymysql
import openpyxl

DB_HOST = '8.163.58.109'
DB_PORT = 3306
DB_USER = 'thai_auto_parts_crm'
DB_PASSWORD = 'tDdY8NX2xJ6HpdHz'
DB_NAME = 'thai_auto_parts_crm'

def connect_db():
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        charset='utf8mb4'
    )

def create_product_table(conn):
    cursor = conn.cursor()
    create_sql = """
CREATE TABLE IF NOT EXISTS `p_product` (
    `id` BIGINT AUTO_INCREMENT PRIMARY KEY COMMENT '主键ID',
    `product_name` VARCHAR(200) NOT NULL COMMENT '产品名称',
    `product_code` VARCHAR(50) COMMENT '产品编码',
    `category` VARCHAR(50) COMMENT '产品类别',
    `brand` VARCHAR(100) COMMENT '品牌',
    `specification` VARCHAR(500) COMMENT '规格',
    `unit_price` DECIMAL(10,2) COMMENT '单价(USD)',
    `stock` INT DEFAULT 0 COMMENT '库存',
    `description` TEXT COMMENT '产品描述',
    `status` VARCHAR(20) DEFAULT 'active' COMMENT '状态:active-在售,inactive-下架',
    `created_at` DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
    `updated_at` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
    
    INDEX `idx_product_name` (`product_name`),
    INDEX `idx_category` (`category`),
    INDEX `idx_status` (`status`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='产品表';
    """
    cursor.execute(create_sql)
    conn.commit()
    print("✅ 产品表创建成功")
    cursor.close()

def import_products(conn):
    cursor = conn.cursor()
    
    cursor.execute("DELETE FROM p_product")
    conn.commit()
    print("已清空现有产品数据")
    
    wb = openpyxl.load_workbook(r'E:\09.document\carparts\产品\产品价格2026-6-25(1).xlsx')
    ws = wb['Sheet1']
    
    insert_count = 0
    for row in range(4, ws.max_row + 1):
        seq = ws.cell(row=row, column=1).value
        if seq is None:
            continue
        
        product_name = ws.cell(row=row, column=2).value
        oe_no = ws.cell(row=row, column=3).value
        weight = ws.cell(row=row, column=4).value
        model = ws.cell(row=row, column=5).value
        dimensions = ws.cell(row=row, column=6).value
        pkg_length = ws.cell(row=row, column=7).value
        pkg_width = ws.cell(row=row, column=8).value
        pkg_height = ws.cell(row=row, column=9).value
        qty_per_pkg = ws.cell(row=row, column=10).value
        car_model = ws.cell(row=row, column=11).value
        production_date = ws.cell(row=row, column=12).value
        price_cny = ws.cell(row=row, column=13).value
        quote_price_cny = ws.cell(row=row, column=14).value
        
        if product_name is None or product_name.strip() == '':
            continue
        
        if isinstance(oe_no, str) and 'VLOOKUP' in oe_no:
            oe_no = None
        
        if isinstance(car_model, str) and 'VLOOKUP' in car_model:
            car_model = None
        
        if isinstance(production_date, str) and 'VLOOKUP' in production_date:
            production_date = None
        
        if price_cny and isinstance(price_cny, (int, float)):
            unit_price_usd = price_cny / 7.2
        else:
            unit_price_usd = 0
        
        category = 'exterior'
        if '扰流板' in product_name:
            category = 'exterior'
        elif '尾管' in product_name:
            category = 'exterior'
        elif '中网' in product_name:
            category = 'exterior'
        elif '前唇' in product_name:
            category = 'exterior'
        elif '侧裙' in product_name:
            category = 'exterior'
        elif '后唇' in product_name:
            category = 'exterior'
        elif '轮眉' in product_name:
            category = 'exterior'
        elif '行李架' in product_name:
            category = 'exterior'
        elif '踏板' in product_name:
            category = 'exterior'
        elif '护板' in product_name:
            category = 'exterior'
        elif '发动机' in product_name:
            category = 'engine'
        elif '刹车' in product_name:
            category = 'brake'
        elif '悬挂' in product_name:
            category = 'suspension'
        elif '电气' in product_name or '灯' in product_name:
            category = 'electrical'
        else:
            category = 'exterior'
        
        spec_parts = []
        if model:
            spec_parts.append(f"型号: {model}")
        if dimensions:
            spec_parts.append(f"尺寸: {dimensions}")
        if weight and isinstance(weight, (int, float)):
            spec_parts.append(f"重量: {weight}g")
        if car_model:
            spec_parts.append(f"适用车型: {car_model}")
        if pkg_length and pkg_width and pkg_height:
            spec_parts.append(f"包装: {pkg_length}x{pkg_width}x{pkg_height}mm")
        if qty_per_pkg:
            spec_parts.append(f"每箱: {qty_per_pkg}件")
        
        specification = '; '.join(spec_parts)
        
        try:
            sql = """
INSERT INTO p_product (product_name, product_code, category, brand, specification, unit_price, stock, description, status)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql, (
                product_name,
                oe_no,
                category,
                'OEM',
                specification,
                round(unit_price_usd, 2),
                100,
                f"{product_name} - 适用车型: {car_model or '通用'}",
                'active'
            ))
            insert_count += 1
            if insert_count % 20 == 0:
                conn.commit()
                print(f"已导入 {insert_count} 条数据...")
        except Exception as e:
            print(f"❌ 导入失败 (行{row}): {product_name} - {e}")
    
    conn.commit()
    print(f"\n✅ 共导入 {insert_count} 条产品数据")
    cursor.close()

def verify_data(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as total FROM p_product")
    result = cursor.fetchone()
    print(f"📊 数据库中共有 {result[0]} 条产品记录")
    
    cursor.execute("SELECT category, COUNT(*) as count FROM p_product GROUP BY category")
    categories = cursor.fetchall()
    print("\n📈 类别分布:")
    for cat, count in categories:
        print(f"  {cat}: {count} 条")
    
    cursor.execute("SELECT MIN(unit_price), MAX(unit_price), AVG(unit_price) FROM p_product")
    price_stats = cursor.fetchone()
    print(f"\n💰 价格统计:")
    print(f"  最低价格: ${price_stats[0]:.2f}")
    print(f"  最高价格: ${price_stats[1]:.2f}")
    print(f"  平均价格: ${price_stats[2]:.2f}")
    
    cursor.close()

if __name__ == '__main__':
    print("=" * 60)
    print("正在连接数据库...")
    print("=" * 60)
    
    try:
        conn = connect_db()
        print("✅ 数据库连接成功")
        
        print("\n" + "=" * 60)
        print("正在创建产品表...")
        print("=" * 60)
        create_product_table(conn)
        
        print("\n" + "=" * 60)
        print("正在导入产品数据...")
        print("=" * 60)
        import_products(conn)
        
        print("\n" + "=" * 60)
        print("数据验证...")
        print("=" * 60)
        verify_data(conn)
        
        conn.close()
        print("\n🎉 所有操作完成！")
        
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()